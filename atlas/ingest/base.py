from abc import ABC, abstractmethod
from pathlib import Path
import polars as pl


class BaseIngester(ABC):
    source_name: str
    category: str

    @property
    def raw_dir(self) -> Path:
        from atlas.config import settings
        d = settings.data_raw_dir / self.category
        d.mkdir(parents=True, exist_ok=True)
        return d

    @abstractmethod
    async def fetch(self) -> None:
        """Download raw data files to self.raw_dir"""

    @abstractmethod
    def parse(self) -> pl.DataFrame:
        """Parse raw files into a clean Polars DataFrame"""

    def save_csv(self, df: pl.DataFrame, name: str) -> Path:
        from atlas.config import settings
        out = settings.data_processed_dir / self.category / f"{name}.csv"
        out.parent.mkdir(parents=True, exist_ok=True)
        df.write_csv(out)
        return out

    def save_excel(self, sheets: dict[str, pl.DataFrame], name: str) -> Path:
        import openpyxl
        from atlas.config import settings
        out = settings.data_processed_dir / self.category / f"{name}.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name, df in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append(df.columns)
            for row in df.iter_rows(named=False):
                ws.append(list(row))
        wb.save(out)
        return out
